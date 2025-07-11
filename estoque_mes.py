import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill
from datetime import datetime

# Configurações iniciais
caminho_pasta = r'C:\Users\win11\OneDrive\Documentos\Planilhas de Custos Médios\2025\Julho'
downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')

# Obter mês e ano do nome da pasta
nome_pasta = os.path.basename(caminho_pasta)
mes_ano = nome_pasta.split(' - ')[-1]  # Pega "Junho 2025"

# Criar nome do arquivo com mês e ano em palavras
arquivo_saida = os.path.join(downloads_path, f'Custos de produtos - {mes_ano}.xlsx')

# Dicionário para armazenar os dados consolidados
dados_consolidados = {}

# Dicionário para armazenar os DataFrames de cada aba individual
abas_individuais = {}

# Lista para armazenar DataFrames da aba Base
dados_base = []

# Lista para armazenar todas as datas encontradas
datas_encontradas = []

# Processar cada arquivo na pasta
for arquivo in os.listdir(caminho_pasta):
    if arquivo.startswith('ev') and (arquivo.endswith('.xlsx') or arquivo.endswith('.csv')):
        # Extrair data do nome do arquivo (assumindo formato evddmmyy)
        data_str = arquivo[2:8]  # Pega os 6 dígitos após 'ev'
        
        try:
            # Converter para objeto de data
            data = datetime.strptime(data_str, '%d%m%y').date()
            data_formatada = data.strftime('%d/%m/%Y')
            datas_encontradas.append(data_formatada)
            
            # Determinar o caminho completo do arquivo
            caminho_arquivo = os.path.join(caminho_pasta, arquivo)
            
            # Ler o arquivo (Excel ou CSV)
            if arquivo.endswith('.xlsx'):
                df = pd.read_excel(caminho_arquivo, skiprows=2)  # Pular 2 linhas de cabeçalho
            else:
                # Para CSV, primeiro detectar o delimitador
                with open(caminho_arquivo, 'r', encoding='latin1') as f:
                    first_line = f.readline()
                    second_line = f.readline()
                    third_line = f.readline()
                
                # Verificar se o cabeçalho está na terceira linha
                if 'PRODUTO' in third_line:
                    # Se o delimitador for tabulação
                    if '\t' in third_line:
                        df = pd.read_csv(caminho_arquivo, skiprows=2, delimiter='\t', encoding='latin1')
                    # Se for ponto e vírgula
                    elif ';' in third_line:
                        df = pd.read_csv(caminho_arquivo, skiprows=2, delimiter=';', encoding='latin1')
                    # Se for vírgula
                    elif ',' in third_line:
                        df = pd.read_csv(caminho_arquivo, skiprows=2, delimiter=',', encoding='latin1')
                    else:
                        # Tentar ler sem especificar delimitador
                        df = pd.read_csv(caminho_arquivo, skiprows=2, encoding='latin1')
                else:
                    # Se 'PRODUTO' não estiver na terceira linha, tentar encontrar
                    df = pd.read_csv(caminho_arquivo, encoding='latin1')
                    # Procurar a linha que contém 'PRODUTO'
                    for i, row in df.iterrows():
                        if 'PRODUTO' in str(row.values):
                            df = pd.read_csv(caminho_arquivo, skiprows=i, encoding='latin1')
                            break
            
            # Renomear colunas para padrão (remover espaços extras)
            df.columns = df.columns.str.strip()
            
            # Adicionar coluna de data para a aba Base
            df['DATA'] = data_formatada
            
            # Armazenar o DataFrame para a aba Base
            dados_base.append(df)
            
            # Armazenar o DataFrame original para escrever depois
            abas_individuais[data_str] = df.drop(columns=['DATA'], errors='ignore')
            
            # Processar dados para a planilha consolidada
            for _, row in df.iterrows():
                # Verificar nomes alternativos das colunas
                produto = row.get('PRODUTO', row.get('Produto', row.get('produto', '')))
                descricao = row.get('DESCRICAO', row.get('Descricao', row.get('descricao', '')))
                grupo = row.get('GRUPO', row.get('Grupo', row.get('grupo', '')))
                custo = row.get('CUSTO', row.get('Custo', row.get('custo', '')))
                
                if produto and pd.notna(produto):
                    if produto not in dados_consolidados:
                        dados_consolidados[produto] = {
                            'DESCRICAO': descricao,
                            'GRUPO': grupo,
                            'CUSTOS': {}
                        }
                    
                    dados_consolidados[produto]['CUSTOS'][data_formatada] = custo
        
        except Exception as e:
            print(f"Erro ao processar o arquivo {arquivo}: {str(e)}")

# Ordenar as datas
datas_encontradas.sort(key=lambda x: datetime.strptime(x, '%d/%m/%Y'))

# Criar DataFrame consolidado
linhas_consolidadas = []
for produto, dados in dados_consolidados.items():
    linha = {
        'PRODUTO': produto,
        'DESCRICAO': dados['DESCRICAO'],
        'GRUPO': dados['GRUPO']
    }
    
    for data in datas_encontradas:
        custo = dados['CUSTOS'].get(data, '')
        # Formatar como moeda se houver valor
        if custo != '' and pd.notna(custo):
            try:
                linha[data] = f'R$ {float(custo):,.2f}'.replace('.', ',').replace(',', '.', 1)
            except:
                linha[data] = custo
        else:
            linha[data] = ''
    
    linhas_consolidadas.append(linha)

df_consolidado = pd.DataFrame(linhas_consolidadas)

# Garantir que todas as colunas de datas estão presentes
for data in datas_encontradas:
    if data not in df_consolidado.columns:
        df_consolidado[data] = ''

# Criar DataFrame para a aba Base
if dados_base:
    df_base = pd.concat(dados_base, ignore_index=True)
    # Reordenar colunas para deixar DATA como primeira coluna
    colunas = ['DATA'] + [col for col in df_base.columns if col != 'DATA']
    df_base = df_base[colunas]
else:
    df_base = pd.DataFrame()

def formatar_como_tabela(worksheet, df, nome_tabela):
    if df.empty:
        return
    
    # Função para converter número de coluna para letra (A, B, ..., Z, AA, AB, etc.)
    def col_to_letter(col):
        letter = ''
        while col > 0:
            col, remainder = divmod(col - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter
    
    # Determinar as dimensões da tabela
    max_row = len(df)
    max_col = len(df.columns)
    
    # Criar a referência no formato correto (ex: "A1:C4")
    start_cell = 'A1'
    end_col = col_to_letter(max_col)
    end_cell = f"{end_col}{max_row + 1}"  # +1 porque a linha 1 é o cabeçalho
    ref = f"{start_cell}:{end_cell}"
    
    try:
        # Criar a tabela
        tab = Table(displayName=nome_tabela, ref=ref)
        
        # Definir um estilo
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        
        # Adicionar a tabela à planilha
        worksheet.add_table(tab)
        
        # Formatar cabeçalho
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Formatar linhas intercaladas
        light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.row % 2 == 0:
                    cell.fill = light_gray
                else:
                    cell.fill = white
    
    except Exception as e:
        print(f"Erro ao formatar tabela {nome_tabela}: {str(e)}")

# Escrever as abas no arquivo Excel
with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
    # Primeiro a aba consolidada
    df_consolidado.to_excel(writer, sheet_name='Consolidado', index=False)
    
    # Depois a aba Base
    if not df_base.empty:
        df_base.to_excel(writer, sheet_name='Base', index=False)
    
    # Por último as abas individuais
    for data_str, df in abas_individuais.items():
        df.to_excel(writer, sheet_name=data_str, index=False)
    
    # Acessar o workbook para formatar as tabelas
    workbook = writer.book
    
    # Formatando cada aba como tabela
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Determinar o DataFrame correspondente
        if sheet_name == 'Consolidado':
            df = df_consolidado
        elif sheet_name == 'Base':
            df = df_base
        else:
            df = abas_individuais.get(sheet_name, pd.DataFrame())
        
        if not df.empty:
            # Nome da tabela (remover caracteres inválidos)
            table_name = f"Table_{sheet_name}".replace(" ", "_").replace("-", "_")
            
            # Chamar função para formatar como tabela
            formatar_como_tabela(worksheet, df, table_name)
        
        # Ajustar a largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

print(f"Processo concluído. Arquivo gerado em: {arquivo_saida}")